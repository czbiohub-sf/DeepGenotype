import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import argparse
import sys
import os

class MyParser(argparse.ArgumentParser):
    def error(self, message):
        sys.stderr.write('error: %s\n' % message)
        self.print_help()
        sys.exit(2)

def parse_args():
    parser = MyParser(description='This script converts the DeepGenotype result csv file to an Excel file with merged headers.')
    parser.add_argument('--path2csv', default="", type=str,
                        help='path to the DeepGenotyp result csv file', metavar='')
    parser.add_argument('--mode', default="", type=str,
                        help='mode, possible values are SNP, INS', metavar='')
    parser.add_argument('--path2csv_consensus', default="", type=str,
                        help='(optional) path to the consensus genotype result csv file', metavar='')

    config = parser.parse_args()
    if len(sys.argv) <= 2:  # print help message if arguments are not valid
        parser.print_help()
        sys.exit(1)
    return config

config = vars(parse_args())

# Define a bold font style
bold_font = Font(bold=True)

# Merge headers based on consecutive identical values
def merge_headers(worksheet, row=1):
    previous_value = None
    start_column = 1

    for col in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(row=row, column=col).value
        if cell_value == previous_value:
            continue
        elif previous_value is not None and cell_value != previous_value:
            if col - start_column > 1:  # There are cells to merge
                worksheet.merge_cells(start_row=row, start_column=start_column, end_row=row, end_column=col-1)
                # Align text to center after merging
                merged_cell = worksheet.cell(row=row, column=start_column)
                merged_cell.alignment = Alignment(horizontal='center')
            start_column = col
        previous_value = cell_value

    # Check for a merge on the last set of columns
    if start_column < worksheet.max_column:
        worksheet.merge_cells(start_row=row, start_column=start_column, end_row=row, end_column=worksheet.max_column)
        merged_cell = worksheet.cell(row=row, column=start_column)
        merged_cell.alignment = Alignment(horizontal='center')

def format_worksheet(ws, mode):
    """Apply standard formatting to a worksheet: merge headers, bold fonts, column widths."""
    merge_headers(ws, row=2)
    merge_headers(ws, row=3)

    if mode == "SNP":
        merge_headers(ws, row=4)
        merge_headers(ws, row=5)

    # Adjust column widths to fit the longest string in each column
    max_length = {}
    for col in ws.columns:
        column = col[0].column_letter  # Get the column letter
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        max_length[column] = max_len + 0.5  # Add a little extra space

    for col, max_len in max_length.items():
        ws.column_dimensions[col].width = max_len

    # manually set the width of the first column
    ws.column_dimensions['A'].width = 27

    # Apply bold font to the first column
    for col in ws.iter_cols(min_col=1, max_col=1, min_row=1, max_row=ws.max_row):
        for cell in col:
            cell.font = bold_font

    # Apply bold font to the first n rows
    if mode == "SNP":
        bold_rows = 7
    else:
        bold_rows = 5
    for row in ws.iter_rows(min_row=1, max_row=bold_rows, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = bold_font


# Load the CSV file into a DataFrame
df = pd.read_csv(config["path2csv"], header = None)

# Save DataFrame to an Excel file initially without any merging
excel_path = config["path2csv"].replace(".csv", ".xlsx")
df.to_excel(excel_path, index=False, engine='openpyxl')

# Load the workbook and select the active worksheet
wb = load_workbook(excel_path)
ws = wb.active

# Determine sheet name based on whether consensus data exists
has_consensus = config["path2csv_consensus"] and os.path.isfile(config["path2csv_consensus"])
if has_consensus:
    ws.title = "Standard Genotypes"
# else leave the default sheet name unchanged for backward compatibility

# Format the standard genotypes sheet
format_worksheet(ws, config["mode"])

# Add consensus genotypes sheet if consensus CSV is provided
if has_consensus:
    df_consensus = pd.read_csv(config["path2csv_consensus"], header=None)
    ws2 = wb.create_sheet(title="Consensus Genotypes")

    # Write consensus DataFrame to the new sheet
    for r_idx, row in enumerate(dataframe_to_rows(df_consensus, index=False, header=False), 1):
        for c_idx, value in enumerate(row, 1):
            ws2.cell(row=r_idx, column=c_idx, value=value)

    # Apply the same formatting
    format_worksheet(ws2, config["mode"])

# Save the modified workbook
wb.save(excel_path)
